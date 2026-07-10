import os
_HERE=os.path.dirname(os.path.abspath(__file__))+os.sep
#!/usr/bin/env python3
"""Export the generated schedule to monthly .xlsx files matching the program's
TYP template, plus a combined workbook."""
import importlib.util, pickle
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
BASE=_HERE
spec=importlib.util.spec_from_file_location("gen",BASE+"gen.py"); gen=importlib.util.module_from_spec(spec); spec.loader.exec_module(gen)
A={date.fromisoformat(k):v for k,v in pickle.load(open(BASE+"assign.pkl","rb"))["assign"].items()}

# ---- full names for footer ----
NAME={
 "MACNEILLE":"Stephen MacNeille","BRONSON":"Isaac Bronson","WISE":"Julien Wise",
 "KENNEDY":"Dean Kennedy","ZAIDI":"Humza Zaidi","OGHENESUME":"Oghenewoma Oghenesume",
 "LI":"Anna Li","MATSUOKA":"Kazune Matsuoka",
 "BUTT":"Aqsa Butt","MULLINS":"Haley Mullins","SAEED":"Usman Saeed","SHETTY":"Kalasha Shetty",
 "VILLANUEVA":"Ricardo Villanueva Gaona","FARZEELA":"Fnu Farzeela","AHN":"Hyojin Ahn",
 "RIVERA":"Angel Maisonet Rivera","GABALLAH":"Bassel Gaballah","METRI":"Nicole Metri","SAEED-S":"Shirin Saeed",
 "AHLUWALIA":"Srishti Ahluwalia","CHIASSON":"Megan Chiasson","VIVEKANANDAN":"Suja Vivekanandan",
 "SALAM":"Muhammed Salam","JUYAL":"Shruti Juyal","KOPP VANUZZI":"Fabio Kopp Vanuzzi",
 "PATEL":"Tirth Pareshbhai Patel","ALMADHOOB":"Mohamed Almadhoob","AHLUWALIA-S":"Saumya Ahluwalia",
 "SANCHEZ-ALMANZAR":"Daniel Sanchez-Almanzar",
}
def grid(lab):   # grid display label (strip disambiguation suffix)
    if lab is None: return ""
    return lab[:-2] if lab.endswith("-S") else lab

# Senior residents (supervisors) — from the roster PDF, for the footer only
SR=[(date(2026,10,12),date(2026,10,25),"Elif Aksoy"),(date(2026,10,26),date(2026,11,8),"Adham Ramadan"),
    (date(2026,11,9),date(2026,12,6),"Ahmed Abouelazaem"),(date(2026,12,7),date(2026,12,20),"Sandresh Sultan"),
    (date(2026,12,21),date(2027,1,17),"Malaika Panchal"),(date(2027,1,28),date(2027,1,31),"Navya Doddareddy"),
    (date(2027,2,1),date(2027,2,14),"Amro Badr"),(date(2027,2,15),date(2027,2,28),"Sana Ali"),
    (date(2027,3,1),date(2027,3,14),"Ranjit Sah"),(date(2027,3,15),date(2027,3,28),"Amro Badr"),
    (date(2027,3,29),date(2027,4,11),"Mohamed Ali"),(date(2027,4,12),date(2027,4,25),"Bera Yildiz"),
    (date(2027,4,26),date(2027,5,9),"Adham Ramadan"),(date(2027,5,10),date(2027,5,23),"Ahmed Abouelazaem"),
    (date(2027,5,24),date(2027,6,6),"Harika Maddisetty"),(date(2027,6,7),date(2027,6,22),"Nishanth Katukuri")]

MONTHS=[(2026,10),(2026,11),(2026,12),(2027,1),(2027,2),(2027,3),(2027,4),(2027,5),(2027,6)]
DAYNAME=["MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY"]
thin=Side(style="thin"); border=Border(thin,thin,thin,thin)
hdrfill=PatternFill("solid",fgColor="D9E1F2"); satfill=PatternFill("solid",fgColor="FCE4D6")
sunfill=PatternFill("solid",fgColor="FFF2CC"); kfill=PatternFill("solid",fgColor="E2EFDA")
bold=Font(bold=True); ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)

def overlaps(blocks, yy, mm):
    ms=date(yy,mm,1); me=gen.month_end(yy,mm); out=[]
    for s,e,lab in blocks:
        if s<=me and e>=ms: out.append((s,e,lab))   # full rotation dates
    return out

def build_sheet(ws, yy, mm):
    ms=date(yy,mm,1); me=min(gen.month_end(yy,mm), gen.SPAN_END)
    ws.column_dimensions["A"].width=13; ws.column_dimensions["B"].width=7
    ws.column_dimensions["C"].width=20; ws.column_dimensions["D"].width=26; ws.column_dimensions["E"].width=22
    ws["A1"]=f"{ms:%B %Y}"; ws["A1"].font=Font(bold=True,size=14)
    hdr=["DAY","DATE","LC: SU-TH 7:00A-6:00P\nFRI 7:00A-7:30P","SC: 7:00A-4:00P",
         "NF: SU-TH 6:00P-8:00A\nFRI 7:30P-9:30A     SAT 24H"]
    for c,h in enumerate(hdr,1):
        cell=ws.cell(2,c,h); cell.font=bold; cell.alignment=ctr; cell.fill=hdrfill; cell.border=border
    r=3; dt=ms
    while dt<=me:
        rec=A[dt]; wd=dt.weekday()
        ws.cell(r,1,DAYNAME[wd]); ws.cell(r,2,dt.day)
        if wd==5:  # Saturday
            ws.cell(r,3,grid(rec["H24"])); ws.cell(r,5,grid(rec["H24"])+"  (24H)")
        elif wd==6:  # Sunday
            ws.cell(r,3,grid(rec["LC"])); ws.cell(r,5,grid(rec["NF"]))
        else:
            ws.cell(r,3,grid(rec["LC"]))
            ws.cell(r,4,", ".join(grid(x) for x in rec["SC"]))
            ws.cell(r,5,grid(rec["NF"]))
        fill=satfill if wd==5 else (sunfill if wd==6 else None)
        for c in range(1,6):
            cell=ws.cell(r,c); cell.border=border; cell.alignment=ctr
            if fill: cell.fill=fill
        r+=1; dt+=timedelta(days=1)
    # ---- footer roster ----
    r+=1
    def section(title, rows):
        nonlocal r
        c=ws.cell(r,3,title); c.font=bold; r+=1
        for nm,rng in rows:
            ws.cell(r,3,nm); ws.cell(r,4,rng); r+=1
        r+=1
    ws.cell(r,3,"INTERNS").font=bold; ws.cell(r,4,"DATES").font=bold; r+=1
    section("LSH", [(NAME[l], f"{ms:%-m/%-d}-{me:%-m/%-d}") for l in gen.LSH_MONTH[(yy,mm)]])
    section("BMC-S / BRIGHTON", [(NAME[l], f"{s:%-m/%-d}-{e:%-m/%-d}") for s,e,l in overlaps(gen.BMC,yy,mm)])
    section("LAHEY", [(NAME[l], f"{s:%-m/%-d}-{e:%-m/%-d}") for s,e,l in overlaps(gen.LAHEY,yy,mm)])
    srrows=[(nm,f"{s:%-m/%-d}-{e:%-m/%-d}") for s,e,nm in SR if s<=me and e>=ms]
    if srrows: section("SENIOR RESIDENT", srrows)
    note=ws.cell(r,3,"SHORT CALL DOESN'T LEAVE BEFORE 4 PM, UNLESS APPROVED BY PD/APDS")
    note.font=Font(bold=True,italic=True)

# one file per month
for yy,mm in MONTHS:
    wb=openpyxl.Workbook(); build_sheet(wb.active,yy,mm); wb.active.title="Sheet1"
    fn=f"{BASE}out/TYP_{date(yy,mm,1):%B_%Y}.xlsx"
    import os; os.makedirs(BASE+"out",exist_ok=True); wb.save(fn)
# combined
wb=openpyxl.Workbook(); wb.remove(wb.active)
for yy,mm in MONTHS:
    ws=wb.create_sheet(f"{date(yy,mm,1):%b %Y}"); build_sheet(ws,yy,mm)
wb.save(f"{BASE}out/TYP_Intern_Schedule_Oct2026-Jun2027.xlsx")
print("Exported", len(MONTHS), "monthly files + combined workbook to", BASE+"out/")
